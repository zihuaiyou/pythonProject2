#
# -*- coding: utf-8 -*-
# # Created by: PyQt5 version 5.15.1 Author:zihuaiyou
# 厚层控底加砂优化软件
import sys
from PyQt5.QtWidgets import *
import mainWin_1_UI
import mainWin_1_UI_child
# import math
from func import *
from scipy.optimize import fsolve
import numpy as np
from openpyxl import Workbook

# # 断裂韧性
# toughness = ""
# # 储层应力
# res_stress = ""
# # 隔层应力
# inter_stress = ""
# # 液体压力
# liq_stress = ""
# # 储层厚度
# res_thick = ""
# # 期望控底高度比
# height_ratio = ""
# # 裂缝长度
# frac_len = ""
# # 粒径 & 粘度
# size_viscosity = ""

str_list = []


# 主界面 厚层控底加砂优化软件
class myMainWindow(QMainWindow, mainWin_1_UI.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    def modify(self):
        # global toughness, res_stress, inter_stress, liq_stress, res_thick, \
        #     height_ratio, frac_len, size_viscosity
        global str_list, str_gradient

        # 读取输入参数(字符串)
        toughness = self.lineEdit_toughness.text() or 1
        res_stress = self.lineEdit_res_stress.text() or 80
        inter_stress = self.lineEdit_inter_stress.text() or 84
        liq_stress = self.lineEdit_liq_stress.text() or 83.2
        res_thick = self.lineEdit_res_thick.text() or 10
        height_ratio = self.lineEdit_height_ratio.text() or 0.2
        frac_len = self.lineEdit_frac_len.text() or 100
        size_viscosity = self.comboBox.currentText()
        # 输入参数列表
        str_list = [toughness, res_stress, inter_stress, liq_stress, res_thick, height_ratio,
                    frac_len, size_viscosity]
        # # 默认值列表
        # default_list = [1, 80, 84, 83.2, 10, 0.2, 100]
        # # 设置默认值
        # for index, value in enumerate(str_list):
        #     if value == "":
        #         str_list[index] = default_list[index]
        # 添加size_viscosity
        # str_list.append(size_viscosity)

        # 1.计算压降梯度
        list_size_viscosity = ['40/70目、30cp', '40/70-70/140、30cp', '70/140目、30cp',
                               '40/70目、200cp', '40/70-70/140、200cp', '70/140目、200cp']
        list_gradient = [1.2, 2.65, 3, 3.3, 7.1, 10]
        dic_gradient = dict(zip(list_size_viscosity, list_gradient))
        # 压降梯度(str)
        str_gradient = str(dic_gradient[str_list[-1]])

        # 2.计算隔板厚度
        # 浮点化输入数据
        [toughness, res_stress, inter_stress, liq_stress, res_thick, height_ratio,
         frac_len] = float_list(str_list[:-1])
        # 隔层高度（float）
        inter_height = res_thick * height_ratio
        #  隔层高度+储层厚度
        THICK = inter_height + res_thick
        # 待求解变量x,隔板厚度
        A = (2 / np.pi) * np.arcsin(res_thick / THICK)
        B = (1 / (np.sqrt(np.pi * THICK))) * toughness / (inter_stress - res_stress)
        C = float(str_gradient) / (np.pi * THICK * (inter_stress - res_stress))
        # D = (3 * THICK - x) / 2
        # E = np.sqrt(2 * THICK - x ** 2)
        F = (THICK ** 2) / 2
        # G = np.arcsin(1 - (x / THICK))
        H = (np.pi / 4) * (THICK ** 2)
        L = (liq_stress - inter_stress) / (res_stress - inter_stress)

        # 公式有误?????
        def func(x):
            return A - B - C * (
                    ((3 * THICK - x) / 2) * np.sqrt(2 * THICK - (x ** 2)) - F * np.arcsin(1 - (x / THICK)) + H) - L

        # 隔板厚度
        # inter_thick = fsolve(func, [4])
        # print(A, B, C, F, H, L)
        # print(inter_thick)


# 子界面 输出隔板设计参数
class myChildWindow(QWidget, mainWin_1_UI_child.Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    # 加砂优化按钮
    def showData(self):
        # 保存为xlsx文件
        sand_size = str_list[-1].split("、")[0]
        file = QFileDialog.getSaveFileName(self, "另存为", ".", "excel(*.xlsx)")
        wb = Workbook()
        ws = wb.active
        # 添加数据文件标题
        ws['A1'] = "加砂优化结果" + " " + "粒径: " + sand_size
        # ws['A2'] = s2
        list_name = ['排量', '静置时间', '砂量', '段塞个数', '砂比']
        ws.append(list_name)
        list1 = [4, 40, '公式', 5, '6%~10%']
        list2 = [4, 20, '公式', 5, '6%~10%']
        list3 = [3, 40, '公式', 5, '6%~10%']
        list4 = [3, 20, '公式', 5, '6%~10%']
        list5 = [2, 40, '公式', 5, '6%~10%']
        list6 = [2, 20, '公式', 5, '6%~10%']
        list7 = [1, 40, '公式', 5, '6%~10%']
        list8 = [1, 20, '公式', 5, '6%~10%']
        ws.append(list1)
        ws.append(list2)
        ws.append(list3)
        ws.append(list4)
        ws.append(list5)
        ws.append(list6)
        ws.append(list7)
        ws.append(list8)
        wb.save(filename=file[0])


if __name__ == '__main__':
    app = QApplication(sys.argv)
    w1 = myMainWindow()
    w1_1 = myChildWindow()
    # 显示主界面1
    w1.show()


    # 显示子界面1
    def show_w1_1():
        w1_1.show()
        # 显示子界面同时文本框显示计算结果值
        w1_1.lineEdit_gradient.setText(str_gradient)


    # 绑定pushbutton
    w1.pushButton.clicked.connect(show_w1_1)

    app.exec_()
